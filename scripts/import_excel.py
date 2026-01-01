import sys
import requests
import openpyxl
from openpyxl import load_workbook

BACKEND_URL = "https://fiscal-control-3.preview.emergentagent.com/api"

def import_from_excel(filepath):
    print("Carregando planilha...")
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['Dados']
    
    # Extrair categorias (linha 2, colunas C-R)
    categorias_data = []
    color_palette = [
        '#FFADAD', '#FFD6A5', '#FDFFB6', '#CAFFBF', '#9BF6FF', '#A0C4FF',
        '#BDB2FF', '#FFC6FF', '#E5E5E5', '#F4A261', '#2A9D8F', '#E9C46A',
        '#E76F51', '#8ECAE6', '#219EBC', '#264653'
    ]
    
    print("\n=== EXTRAINDO CATEGORIAS ===")
    for col_idx in range(3, 19):  # C (3) até R (18) - incluindo Água
        cat_name = ws.cell(row=2, column=col_idx).value
        if cat_name:
            cat_name_clean = cat_name.replace('\n', ' ').strip()
            
            # Extrair dia de vencimento do nome
            due_day = None
            if 'Dia' in cat_name_clean:
                import re
                match = re.search(r'Dia\s+(\d+)', cat_name_clean)
                if match:
                    due_day = int(match.group(1))
            
            categorias_data.append({
                'name': cat_name_clean,
                'due_day': due_day,
                'color': color_palette[(col_idx - 3) % len(color_palette)],
                'order': col_idx - 2,
                'col_idx': col_idx
            })
            print(f"  {col_idx - 2}. {cat_name_clean} (Dia {due_day if due_day else 'N/A'})")
    
    # Limpar banco de dados
    print("\n=== LIMPANDO BANCO DE DADOS ===")
    try:
        # Buscar categorias existentes
        response = requests.get(f"{BACKEND_URL}/categories")
        if response.status_code == 200:
            existing_cats = response.json()
            for cat in existing_cats:
                requests.delete(f"{BACKEND_URL}/categories/{cat['id']}")
            print(f"  Removidas {len(existing_cats)} categorias existentes")
    except Exception as e:
        print(f"  Erro ao limpar: {e}")
    
    # Criar categorias no backend
    print("\n=== CRIANDO CATEGORIAS NO BACKEND ===")
    categoria_map = {}
    
    for cat_data in categorias_data:
        payload = {
            'name': cat_data['name'],
            'due_day': cat_data['due_day'],
            'color': cat_data['color'],
            'order': cat_data['order']
        }
        
        response = requests.post(f"{BACKEND_URL}/categories", json=payload)
        if response.status_code == 200:
            cat_created = response.json()
            categoria_map[cat_data['col_idx']] = cat_created['id']
            print(f"  ✓ {cat_data['name']}")
        else:
            print(f"  ✗ Erro ao criar {cat_data['name']}: {response.text}")
    
    # Extrair e importar transações (valores de cada mês)
    print("\n=== IMPORTANDO VALORES MENSAIS ===")
    meses_nome = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
    
    transactions_created = 0
    for row_idx in range(3, 15):  # Linhas 3-14 (jan-dez)
        mes_nome = ws.cell(row=row_idx, column=2).value
        mes_num = row_idx - 2  # jan=1, fev=2, etc.
        
        print(f"\n  {mes_nome.upper()} (mês {mes_num}):")
        
        for col_idx, cat_id in categoria_map.items():
            valor = ws.cell(row=row_idx, column=col_idx).value
            if valor is None:
                valor = 0
            valor = float(valor)
            
            # Criar transação (usando mesmo valor para planned e actual)
            payload = {
                'category_id': cat_id,
                'month': mes_num,
                'year': 2026,
                'planned_value': valor,
                'actual_value': valor,
                'notes': 'Importado da planilha Excel'
            }
            
            response = requests.post(f"{BACKEND_URL}/transactions", json=payload)
            if response.status_code == 200:
                transactions_created += 1
            else:
                print(f"    ✗ Erro ao criar transação: {response.text}")
        
        print(f"    ✓ {len(categoria_map)} transações criadas")
    
    print(f"\n{'='*60}")
    print(f"IMPORTAÇÃO CONCLUÍDA!")
    print(f"  • {len(categorias_data)} categorias criadas")
    print(f"  • {transactions_created} transações criadas")
    print(f"{'='*60}")

if __name__ == "__main__":
    import_from_excel("/tmp/orcamento2026_new.xlsx")
