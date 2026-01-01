import requests
import openpyxl

BACKEND_URL = "https://fiscal-control-3.preview.emergentagent.com/api"

def import_incomes():
    print("="*70)
    print("IMPORTANDO RECEITAS DA PLANILHA")
    print("="*70)
    
    wb = openpyxl.load_workbook('/tmp/orcamento2026_new.xlsx', data_only=True, read_only=True)
    ws = wb['Dados']
    
    # Limpar receitas existentes
    incomes = requests.get(f"{BACKEND_URL}/incomes?year=2026").json()
    for income in incomes:
        requests.delete(f"{BACKEND_URL}/incomes/{income['id']}")
    print(f"✓ Removidas {len(incomes)} receitas existentes")
    
    # Importar receitas (linhas 19-30)
    created = 0
    for row_idx in range(19, 31):
        mes_num = row_idx - 18
        aposentadoria = ws.cell(row=row_idx, column=4).value or 0
        salario = ws.cell(row=row_idx, column=8).value or 0
        recursos_externos = ws.cell(row=row_idx, column=9).value or 0
        
        payload = {
            'month': mes_num,
            'year': 2026,
            'aposentadoria': round(float(aposentadoria), 2),
            'salario': round(float(salario), 2),
            'recursos_externos': round(float(recursos_externos), 2),
            'notes': 'Importado da planilha Excel'
        }
        
        response = requests.post(f"{BACKEND_URL}/incomes", json=payload)
        if response.status_code == 200:
            created += 1
            data = response.json()
            total = data['aposentadoria'] + data['salario'] + data['recursos_externos']
            print(f"✓ Mês {mes_num:2d}: R$ {total:,.2f}")
    
    print(f"\n{'='*70}")
    print(f"✓ {created} receitas importadas com sucesso")
    print(f"{'='*70}")

def reset_actual_values():
    print("\n" + "="*70)
    print("ZERANDO VALORES REALIZADOS")
    print("="*70)
    
    response = requests.post(f"{BACKEND_URL}/reset-actual-values")
    if response.status_code == 200:
        data = response.json()
        print(f"✓ {data['modified_count']} transações atualizadas")
        print("✓ Todos os valores 'realizados' foram zerados")
        print("✓ Valores 'planejados' mantidos intactos")
    else:
        print(f"✗ Erro: {response.text}")
    
    print(f"{'='*70}")

if __name__ == "__main__":
    import_incomes()
    reset_actual_values()
